import bs4
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import openpyxl
import json


class OddsPortal:

    def __init__(self):

        self.__config_path = "config.json"

        self.__sport = ""
        self.__country = ""
        self.__league = ""
        self.__year = 0

        self.__ws_row = 2

        self.__chrome_driver = webdriver.Chrome("C:\\Users\\pcost\\chromedriver_win32\\chromedriver.exe")
        self.__delay = 60

        self.__wb = openpyxl.Workbook()
        self.__wb_name = ""
        self.__ws = self.__wb.active

    def close(self):
        self.__chrome_driver.close()

    def load(self):

        with open(self.__config_path) as f:
            config = json.load(f)
            self.__sport = config["sport"]
            self.__country = config["country"]
            self.__league = config["league"]
            self.__year = config["year"]
            self.__ws.title = str(config["year"])
            self.__wb_name = f'{config["sport"]}_{config["league"]}_{config["year"]}_{config["year"]+1}.xlsx'
            self.__create_ws_columns()

    def __create_ws_columns(self):

        columns = [
            "HT", "AT",
            "DAY", "MONTH", "YEAR",
            "HG", "AG",
            "HO", "DO", "AO",
            "FTO 0.5", "FTU 0.5", "FTO 1.5", "FTU 1.5", "FTO 2.5", "FTU 2.5", "FTO 3.5", "FTU 3.5",
        ]

        col = 1
        for name in columns:
            if not (name == "DO" and self.__sport == "basketball"):  # basketball draws odds are not relevant
                self.__ws.cell(row=1, column=col, value=name)
                col += 1

        self.__wb.save(self.__wb_name)

    def scrap_season(self):

        page = 1
        while True:
            if self.__scrap_page(page) == "END_OF_PAGES":
                break
            page += 1

    def __scrap_page(self, page):
        url = "https://www.oddsportal.com/{}/{}/{}-{}-{}/results/#/page/{}/".format(self.__sport, self.__country, self.__league, self.__year, self.__year + 1, page)
        self.__chrome_driver.get(url)
        try:
            myElem = WebDriverWait(self.__chrome_driver, self.__delay).until(EC.presence_of_element_located((By.CLASS_NAME, 'table-main')))
        except TimeoutException:
            print("Loading page took too much time!")

        soup = bs4.BeautifulSoup(self.__chrome_driver.page_source, "html.parser")
        table = soup.find("table", id="tournamentTable")

        if self.__is_end_of_pages(table):
            return "END_OF_PAGES"

        for tr in table.findAll("tr", {"class": ["odd deactivate", "deactivate"]}):
            match_right_url = tr.find("td", {"class": "name table-participant"}).find("a")["href"]
            match_url = "https://www.oddsportal.com" + match_right_url
            self.__scrap_match(match_url)

        return "SUCCESS"

    @staticmethod
    def __is_end_of_pages(table):
        return table.find(id="emptyMsg") is not None

    def __scrap_match(self, game_url):

        self.__chrome_driver.get(game_url)
        try:
            myElem = WebDriverWait(self.__chrome_driver, self.__delay).until(EC.presence_of_element_located((By.CLASS_NAME, 'result')))
            myElem = WebDriverWait(self.__chrome_driver, self.__delay).until(EC.presence_of_element_located((By.CLASS_NAME, 'table-container')))
        except TimeoutException:
            print("Loading match page took too much time!")

        soup = bs4.BeautifulSoup(self.__chrome_driver.page_source, "html.parser")

        # get team names
        teams_names = soup.find("div", id="col-content").find("h1").get_text().split(" - ")
        home_team = teams_names[0]
        self.__ws.cell(row=self.__ws_row, column=1, value=home_team)
        away_team = teams_names[1]
        self.__ws.cell(row=self.__ws_row, column=2, value=away_team)

        # get match date
        date_list = soup.find("div", id="col-content").find("p", class_="date").get_text().replace(",", " ").split()
        try:
            day = int(date_list[1])
            self.__ws.cell(row=self.__ws_row, column=3, value=day)

            month = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"].index(date_list[2]) + 1
            self.__ws.cell(row=self.__ws_row, column=4, value=month)

            year = int(date_list[3])
            self.__ws.cell(row=self.__ws_row, column=5, value=year)
        except ValueError:
            print("Failed scrapping date")
            self.__ws.cell(row=self.__ws_row, column=3, value=-1)
            self.__ws.cell(row=self.__ws_row, column=4, value=-1)
            self.__ws.cell(row=self.__ws_row, column=5, value=-1)

        # get number goals
        try:
            final_result_soup = soup.find("div", id="event-status").find("p")

            if final_result_soup.find("span").get_text() == "Canceled":
                print("Game Canceled!")
                return

            final_result_str = final_result_soup.find("strong").get_text()

            if 'OT' in final_result_str:
                goals_str = final_result_str[final_result_str.index('(')+1:final_result_str.index(')')]
            else:
                goals_str = final_result_str

            home_goals, away_goals = tuple(goals_str.split(":"))
            self.__ws.cell(row=self.__ws_row, column=6, value=home_goals)
            self.__ws.cell(row=self.__ws_row, column=7, value=away_goals)
        except ValueError:
            print("Failed scrapping goals")
            self.__ws.cell(row=self.__ws_row, column=6, value=-1)
            self.__ws.cell(row=self.__ws_row, column=7, value=-1)

        # get odds
        if self.__sport == "soccer":
            # get result odds
            self.__scrap_soccer_odds(soup)

            # get final time goals odds
            self.__scrap_final_time_goals_odds()

        elif self.__sport == "basketball":
            self.__scrap_basketball_odds(soup)
        else:
            print("Invalid sport")

        self.__ws_row += 1
        print("Match: {}".format(self.__ws_row))
        self.__wb.save(self.__wb_name)

    def __scrap_soccer_odds(self, soup):
        odds = self.__scrap_odds(soup)

        home_odd = odds[0]
        self.__ws.cell(row=self.__ws_row, column=8, value=home_odd)

        draw_odd = odds[1]
        self.__ws.cell(row=self.__ws_row, column=9, value=draw_odd)

        away_odd = odds[2]
        self.__ws.cell(row=self.__ws_row, column=10, value=away_odd)

    def __scrap_basketball_odds(self, soup):
        odds = self.__scrap_odds(soup)

        home_odd = odds[0]
        self.__ws.cell(row=self.__ws_row, column=8, value=home_odd)

        away_odd = odds[1]
        self.__ws.cell(row=self.__ws_row, column=9, value=away_odd)

    def __scrap_final_time_goals_odds(self):
        self.__chrome_driver.find_element(By.XPATH, '//*[@title="Over/Under"]').click()
        try:
            myElem = WebDriverWait(self.__chrome_driver, self.__delay).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'table-container')))
            print("Odds goals full time page is ready!")

        except TimeoutException:
            print("Loading odds goals full time page took too much time!")
        soup = bs4.BeautifulSoup(self.__chrome_driver.page_source, "html.parser")
        self.__scrap_odds_goals_part(soup)

    def __scrap_odds_goals_part(self, soup):
        odds_data_list_soup = soup.find("div", id="odds-data-table").findAll("div", class_="table-container")

        for div in odds_data_list_soup:
            if "Over/Under +0.5" in str(div):
                self.__scrap_odds_over_under(div, 11)
            elif "Over/Under +1.5" in str(div):
                self.__scrap_odds_over_under(div, 13)
            elif "Over/Under +2.5" in str(div):
                self.__scrap_odds_over_under(div, 15)
            elif "Over/Under +3.5" in str(div):
                self.__scrap_odds_over_under(div, 17)

    def __scrap_odds_over_under(self, odds_goals_div, col):
        odds_over_under_span_list = odds_goals_div.findAll("span", class_="avg chunk-odd nowrp")
        try:
            over = float(odds_over_under_span_list[1].find("a").get_text())
            self.__ws.cell(row=self.__ws_row, column=col, value=over)
            under = float(odds_over_under_span_list[0].find("a").get_text())
            self.__ws.cell(row=self.__ws_row, column=col+1, value=under)
        except ValueError:
            print("Odds over/under not real!")

    @staticmethod
    def __scrap_odds(soup):
        result_odds_soup_list = soup.find("div", id="odds-data-table").find("div", {"class": "table-container"}).find("table").find("tfoot").find("tr", {"class": "aver"}).findAll("td", {"class": "right"})
        result_odds_list = []
        for odd_soup in result_odds_soup_list:
            try:
                result_odds_list.append(float(odd_soup.get_text()))
            except (ValueError, AttributeError):
                print("Failed scrapping odds average")
                result_odds_list.append(-1)
        return result_odds_list

